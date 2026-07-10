# -*- coding: utf-8 -*-
"""每月調價建議腳本。

用法：
    python run_monthly.py

資料來源：
    銷售/品牌/類別/成本：BigQuery igogo-sales-dw.sales（v_sales 全通路 + dim_item）
    建議售價：ECOUNT API 的 NO_USER1（見 ecount.py）
    現有庫存（選用）：data/cost/ 的庫存成本報表；有才啟用庫存出清判斷，
                      並取其移動平均成本（比 DW in_price 準）

產出：
    output/調價指令表_YYYYMM.xlsx、output/dashboard.html
    history/recommendations.csv（供下月急煞檢查）
"""
import sys, io, re, json, glob, os
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl.styles.colors as _oc
_orig_rgb = _oc.RGB.__set__
def _rgb_patch(self, inst, v):
    if v is not None and not re.match(r'^([A-Fa-f0-9]{8}|[A-Fa-f0-9]{6})$', str(v)):
        v = '00000000'
    _orig_rgb(self, inst, v)
_oc.RGB.__set__ = _rgb_patch

import pandas as pd
import numpy as np

BASE = os.path.dirname(os.path.abspath(__file__))
CFG = json.load(open(os.path.join(BASE, 'config.json'), encoding='utf-8'))

def fit_panel(g):
    """lnQ ~ lnP + 月序 + 月份dummies + 型號dummies → (E, se)"""
    g = g.copy()
    g['lnQ'] = np.log(g['qty'])
    g['lnP'] = np.log(g['price'])
    g['t'] = (g['month'] - g['month'].min()).map(lambda x: x.n)
    mon = pd.get_dummies(g['month'].dt.month, prefix='m', drop_first=True).astype(float)
    mod = pd.get_dummies(g['model'], drop_first=True).astype(float)
    X = pd.concat([g['lnP'].reset_index(drop=True), g['t'].reset_index(drop=True),
                   mon.reset_index(drop=True), mod.reset_index(drop=True)], axis=1)
    X.insert(0, 'const', 1.0)
    y = g['lnQ'].values
    if len(g) <= X.shape[1] + 3:
        return None, None
    beta, *_ = np.linalg.lstsq(X.values, y, rcond=None)
    resid = y - X.values @ beta
    dof = len(g) - X.shape[1]
    try:
        cov = np.linalg.inv(X.values.T @ X.values) * (resid @ resid) / dof
    except np.linalg.LinAlgError:
        return None, None
    i = list(X.columns).index('lnP')
    return beta[i], np.sqrt(cov[i, i])


def main():
    import datasource
    mo, dim, allq = datasource.load_monthly(CFG)
    mo = mo.rename(columns={'ym': 'month'})
    allq = allq.rename(columns={'ym': 'month'})
    period = f"{mo['month'].min()} ~ {mo['month'].max()}"
    print(f'BigQuery：定價用電商 {len(mo)} 月列（{mo["item_code"].nunique()} 品項），'
          f'庫存消耗用全通路 {len(allq)} 月列，{period}')

    # 異常過濾：BQ 已排除 amt_anomaly；再剔月均價低於型號中位60%者（沖銷/特殊結價）
    med = mo.groupby('model')['price'].transform('median')
    n_out = int((mo['price'] < 0.6 * med).sum())
    if n_out:
        print(f'剔除異常低價月列：{n_out} 筆（<型號中位價60%）')
    mo = mo[mo['price'] >= 0.6 * med].copy()

    # ---- 彈性估計（品類=品牌x類別，其次品牌，其次預設）----
    ok = mo.groupby('model').filter(
        lambda g: len(g) >= 8 and g['price'].mean() > 0
        and g['price'].std() / g['price'].mean() >= 0.03)

    def fit_ok(g):
        if len(g) < CFG['brand_min_obs'] or g['model'].nunique() < CFG['brand_min_models']:
            return None
        e, se = fit_panel(g)
        if e is not None and e <= -1.0 and se is not None and abs(e / se) >= 2:
            return e, se
        return None

    seg_E = {s: r for s, g in ok.groupby('seg') if (r := fit_ok(g))}
    brand_E = {b: r for b, g in ok.groupby('brand') if (r := fit_ok(g))}
    print(f'估出品類彈性 {len(seg_E)} 個、品牌彈性 {len(brand_E)} 個')
    E_default = CFG['default_elasticity']

    def get_E(brand, seg):
        if seg in seg_E:
            return seg_E[seg][0], '品類'
        if brand in brand_E:
            return brand_E[brand][0], '品牌'
        return E_default, '預設'

    # ---- ECOUNT API：建議售價 NO_USER1 ＋ 即時庫存 ----
    retail_map, stock_map = {}, {}
    try:
        import ecount
        master, src = ecount.fetch_master(CFG['secrets_env'])
        retail_map = {c: v['retail'] for c, v in master.items() if v['retail'] > 0}
        print(f'ECOUNT 建議售價（{src}）：{len(retail_map)} 個品項')
    except Exception as e:
        print(f'!! ECOUNT 抓建議售價失敗，改用供貨價推估：{e}')
    # 庫存優先讀 BigQuery v_inventory（每日快照、有歷史）；失敗才回退 ECOUNT 即時
    try:
        inv, snap = datasource.load_inventory(CFG)
        stock_map = {c: q for c, q in inv.items() if q and q > 0}
        print(f'BigQuery 庫存（v_inventory {snap}）：{len(stock_map)} 個品項')
    except Exception as e:
        print(f'!! BigQuery 庫存讀取失敗，回退 ECOUNT 即時：{e}')
        try:
            import ecount
            stock_map, ssrc = ecount.fetch_stock(CFG['secrets_env'])
            stock_map = {c: q for c, q in stock_map.items() if q and q > 0}
            print(f'ECOUNT 即時庫存（{ssrc}）：{len(stock_map)} 個品項')
        except Exception as e2:
            print(f'!! ECOUNT 庫存也失敗：{e2}')

    # ---- 成本：庫存成本報表移動平均優先（選用），否則 DW in_price ----
    dim_cost = {r['code']: r for _, r in dim.iterrows()}
    cost_report = {}
    cost_files = glob.glob(os.path.join(BASE, 'data', 'cost', '*.xlsx'))
    if cost_files:
        cf = max(cost_files, key=os.path.getmtime)
        cr = pd.read_excel(cf, header=1)
        cr.columns = (['code', 'name', 'stock_qty', 'unit_cost', 'stock_value'] + list(cr.columns))[:len(cr.columns)]
        cr = cr.dropna(subset=['code'])
        cost_report = dict(zip(cr['code'], cr['unit_cost']))
        if not stock_map:   # API 失敗才回退報表庫存
            stock_map = {k: v for k, v in zip(cr['code'], cr['stock_qty']) if pd.notna(v) and v > 0}
        print(f'庫存成本報表：{os.path.basename(cf)}（{len(cr)} 品項，提供移動平均成本）')
    else:
        print('（data/cost 無報表：成本改用 DW in_price）')

    # ---- 每品項近期指標（全通路）----
    rr_m = max(1, round(CFG['run_rate_days'] / 30.4))
    p_m = max(1, round(CFG['recent_price_days'] / 30.4))
    maxm = mo['month'].max()
    # 供貨均價 = 電商近期營收加權；月均銷量 = 全通路（庫存消耗速度）
    p_recent = mo[mo['month'] > maxm - p_m].groupby('item_code').apply(
        lambda g: g['rev'].sum() / g['qty'].sum(), include_groups=False).to_dict()
    rr = (allq[allq['month'] > maxm - rr_m].groupby('item_code')['qty'].sum() / rr_m).to_dict()
    mos = mo.sort_values('month')
    attr = mos.groupby('item_code').agg(
        name=('item_name', 'last'), brand=('brand', 'last'), seg=('seg', 'last')).to_dict('index')
    last_p = mos.groupby('item_code').apply(
        lambda g: g['rev'].tail(3).sum() / g['qty'].tail(3).sum(), include_groups=False).to_dict()

    front, back = CFG['platform_front_margin'], CFG['platform_back_margin']
    hold, mstep = CFG['hold_band'], CFG['max_step']

    # ---- 競品錨定（PCHome+MOMO 耳機快照；漲價天花板）----
    import competitor
    anchor = competitor.build_anchor(CFG)
    ceil_f = CFG.get('comp_ceiling_factor', 1.05)
    if anchor:
        print('競品錨定：已載入耳機競品快照，漲價將受同形態市場中位煞車')
    else:
        print('（無競品快照：跳過漲價天花板）')

    # ---- 逐 SKU 建議（只對有電商銷售的品項；月均銷量取全通路）----
    rows = []
    for code in last_p:
        C = cost_report.get(code)
        if C is None or pd.isna(C) or C <= 0:
            d = dim_cost.get(code)
            C = float(d['unit_cost']) if d is not None and d['unit_cost'] else np.nan
        if pd.isna(C) or C <= 0:
            continue
        P = p_recent.get(code) or last_p.get(code)
        if not P or P <= 0:
            continue
        at = attr.get(code, {})
        brand, seg, name = at.get('brand', ''), at.get('seg', ''), at.get('name', code)
        rate = float(rr.get(code, 0))
        stock = stock_map.get(code, np.nan)
        moS = stock / rate if (rate > 0 and pd.notna(stock)) else np.nan
        E, esrc = get_E(brand, seg)
        a = abs(E)

        # 建議售價（定價）：優先用 ECOUNT NO_USER1 官方定價，缺則用供貨價推估
        R = retail_map.get(code, 0)
        if R and R > P:
            r_src = 'ERP'
        else:
            R = P / (1 - front)
            r_src = '推估'
        f = P / R                       # 供貨價/定價 = 1 − 實際前毛利（逐品項）
        net = P - back * R              # 淨收入 = 供貨價 − 後段維運(定價 x 後毛利)
        denom = f - back                # 定價每 +1，淨收入 + denom
        margin = (net - C) / net if net > 0 else -9
        target_m = min(1 / a, CFG['max_target_margin'])

        # 讓淨毛利達標所需的定價（denom>0 才有解）
        if denom > 0:
            R_star = C / ((1 - target_m) * denom)
            R_floor = C / ((1 - CFG['min_margin']) * denom)
        else:
            R_star = R_floor = R        # 結構性虧損，靠定價救不回
        cap = mstep if esrc != '預設' else mstep / 2   # 預設彈性=假設值，步幅減半
        step = float(np.clip(R_star / R - 1, -cap, cap))

        if margin < 0:
            act, why = '檢視-負毛利', f'淨收入{net:.0f}低於成本{C:.0f}，先查售價或成本是否正確，再決定停售/漲價/認賠出清'
        elif CFG['use_stock_logic'] and pd.notna(moS) and moS > CFG['overstock_months'] and step >= 0:
            if margin > target_m:
                act, why = '降價出清', f'庫存{moS:.0f}個月過深，毛利{margin:.0%}高於目標{target_m:.0%}，有降價空間'
                step = -min(cap, margin - target_m)
            else:
                act, why = '贈品出清', f'庫存{moS:.0f}個月過深但毛利{margin:.0%}已低於目標{target_m:.0%}，降價不划算，用贈品/組合'
                step = 0.0
        elif abs(step) < hold:
            act, why = '不動', f'毛利{margin:.0%}≈目標{target_m:.0%}，已在最佳價附近'
        elif step > 0:
            act, why = '漲價', f'毛利{margin:.0%}<目標{target_m:.0%}（彈性{a:.1f}），每月最多+{cap:.0%}逐步走高'
        else:
            act, why = '降價', f'毛利{margin:.0%}>目標{target_m:.0%}（彈性{a:.1f}），降價換量可增利'

        # 定調後統一換算：定價取整到10元，供貨價依實際折數 f 同步移動
        new_R = round(R * (1 + step) / 10) * 10 if abs(step) > 1e-9 else round(R)

        # 競品錨定：漲價不得超過同形態市場中位 x 溢價倍數
        mkt_med = mkt_low = mkt_n = None
        a_info = anchor(competitor.form_of(name), R) if anchor else None
        if a_info:
            mkt_med, mkt_low, mkt_n = round(a_info['median']), round(a_info['low']), a_info['n']
            ceiling = a_info['median'] * ceil_f
            if act == '漲價' and new_R > ceiling:
                capped = max(round(ceiling / 10) * 10, round(R))   # 不低於現價
                if capped <= R:      # 現價已達/超過競品天花板 → 別漲
                    act, new_R = '不動', round(R)
                    why = f'毛利{margin:.0%}<目標但已達競品天花板（同形態市場中位{mkt_med}），漲價會失去價格競爭力'
                else:
                    new_R = capped
                    why = f'漲價至競品天花板（同形態市場中位{mkt_med}x{ceil_f:.2f}）；毛利{margin:.0%}<目標但不宜漲過市場'

        step = new_R / R - 1
        new_P = f * new_R
        pred = (1 + step) ** E - 1 if abs(step) > 1e-9 else 0.0
        vs_mkt = round(R / mkt_med - 1, 3) if mkt_med else None

        rows.append(dict(品項編碼=code, 品項名=name, 品牌=brand,
                         彈性=round(E, 2), 彈性來源=esrc,
                         庫存量=(round(stock) if pd.notna(stock) else None), 月均銷量=round(rate, 1),
                         庫存月數=(round(moS, 1) if pd.notna(moS) else None), 目前售價=round(R), 售價來源=r_src,
                         供貨均價=round(P), 淨收入=round(net), 成本=round(C, 1),
                         真實毛利率=round(margin, 3), 目標毛利率=round(target_m, 3),
                         市場中位=mkt_med, vs市場=vs_mkt, 競品數=mkt_n,
                         方向=act, 建議售價=int(new_R), 建議供貨價=round(new_P),
                         調幅=round(step, 3), 預期銷量變化=round(pred, 3), 理由=why))

    out = pd.DataFrame(rows)
    order = {'檢視-負毛利': 0, '漲價': 1, '降價': 2, '降價出清': 3, '贈品出清': 4, '不動': 5}
    out['_o'] = out['方向'].map(order)
    out['_rev'] = out['月均銷量'] * out['供貨均價']   # 依營收規模排序（庫存量可能缺）
    out = out.sort_values(['_o', '_rev'], ascending=[True, False]).drop(columns=['_o', '_rev'])

    # ---- 急煞檢查（跟上月建議比對）----
    hist_path = os.path.join(BASE, 'history', 'recommendations.csv')
    alerts = []
    cur_tag = datetime.now().strftime('%Y%m')
    if os.path.exists(hist_path):
        hist = pd.read_csv(hist_path, encoding='utf-8-sig')
        hist = hist[hist['run'].astype(str) < cur_tag]   # 同月重跑不觸發急煞
        last_run = hist['run'].max() if len(hist) else None
        prev = hist[hist['run'] == last_run] if last_run else pd.DataFrame()
        this_m = mo[mo['month'] == maxm].set_index('item_code')['qty']
        prev_m = mo[mo['month'] == maxm - 1].set_index('item_code')['qty']
        for _, h in prev.iterrows():
            a_now = float(this_m.get(h['品項編碼'], 0))
            a_prev = float(prev_m.get(h['品項編碼'], 0))
            if a_prev >= 5 and abs(h['調幅']) >= 0.03:
                actual = a_now / a_prev - 1
                if actual < h['預期銷量變化'] - 0.5:
                    alerts.append(f"{h['品項編碼']} {h['品項名']}：預期{h['預期銷量變化']:+.0%}實際{actual:+.0%}，建議退回原價")
        if alerts:
            print('\n!! 急煞警報：')
            for a_ in alerts:
                print('  ', a_)

    run_tag = datetime.now().strftime('%Y%m')
    out.insert(0, 'run', run_tag)
    if os.path.exists(hist_path):
        old = pd.read_csv(hist_path, encoding='utf-8-sig')
        old = old[old['run'].astype(str) != run_tag]     # 同月重跑覆蓋
        pd.concat([old, out], ignore_index=True).to_csv(hist_path, index=False, encoding='utf-8-sig')
    else:
        out.to_csv(hist_path, index=False, encoding='utf-8-sig')

    # ---- 輸出 Excel ----
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = '調價指令'
    BOLD = Font(name='Arial', bold=True)
    HDR = PatternFill('solid', start_color='D9D9D9')
    FILLS = {'檢視-負毛利': 'F4CCCC', '漲價': 'D9EAD3', '降價': 'FCE5CD',
             '降價出清': 'FCE5CD', '贈品出清': 'FFF2CC'}
    cols = [c for c in out.columns if c != 'run']
    for j, h in enumerate(cols, 1):
        cell = ws.cell(1, j, h)
        cell.font = BOLD; cell.fill = HDR
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    for i, (_, r) in enumerate(out.iterrows(), 2):
        for j, h in enumerate(cols, 1):
            cell = ws.cell(i, j, r[h])
            cell.font = Font(name='Arial')
            if h == '方向' and r[h] in FILLS:
                cell.fill = PatternFill('solid', start_color=FILLS[r[h]])
            if h in ('真實毛利率', '目標毛利率', '調幅', '預期銷量變化', 'vs市場'):
                cell.number_format = '0.0%'
            if h in ('目前售價', '供貨均價', '淨收入', '成本', '建議售價', '建議供貨價', '市場中位'):
                cell.number_format = '#,##0'
    widths = [11, 40, 10, 7, 9, 8, 9, 9, 9, 8, 9, 9, 9, 9, 9, 9, 8, 7, 11, 10, 10, 7, 10, 60]
    for j, w in enumerate(widths[:len(cols)], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{len(out)+1}'

    ws2 = wb.create_sheet('本次參數')
    n_erp = int((out['售價來源'] == 'ERP').sum())
    params = [('資料來源', f'BigQuery v_sales 全通路（{period}）'),
              ('建議售價來源', f"ECOUNT API 官方定價 {n_erp} 個 / 推估 {len(out)-n_erp} 個"),
              ('預設彈性(品牌估不出時)', f'{E_default:.2f}'),
              ('平台前毛利', f"{front:.0%}（僅缺定價時推估用；有定價則逐品項實算）"),
              ('平台後毛利', f"{back:.0%}（定價 x 此比例＝後段維運）"),
              ('單月最大調幅', f'{mstep:.0%}'), ('滯銷門檻', f"{CFG['overstock_months']} 個月"),
              ('最低毛利地板', f"{CFG['min_margin']:.0%}")]
    for s, (e, se) in sorted(seg_E.items()):
        params.append((f'{s} 彈性', f'{e:.2f} ± {se:.2f}'))
    for b, (e, se) in sorted(brand_E.items()):
        params.append((f'{b} 整體彈性', f'{e:.2f} ± {se:.2f}'))
    for i, (a_, b_) in enumerate(params, 1):
        ws2.cell(i, 1, a_).font = BOLD
        ws2.cell(i, 2, b_).font = Font(name='Arial')
    ws2.column_dimensions['A'].width = 22; ws2.column_dimensions['B'].width = 40

    out_path = os.path.join(BASE, 'output', f'調價指令表_{run_tag}.xlsx')
    wb.save(out_path)

    # ---- 儀表板 ----
    import dashboard
    elas = {s: f'{e:.2f} ± {se:.2f}' for s, (e, se) in sorted(seg_E.items())}
    for b, (e, se) in sorted(brand_E.items()):
        if not any(key.startswith(b + '·') for key in elas):
            elas[b] = f'{e:.2f} ± {se:.2f}'
    meta = dict(run=run_tag, period=period,
                front=round(front * 100), back=round(back * 100),
                erp_retail=int((out['售價來源'] == 'ERP').sum()), n=len(out),
                comp_anchored=bool(anchor),
                max_step=round(mstep * 100), default_e=E_default, elasticity=elas)
    recs = out.drop(columns=['run']).replace({np.nan: None}).to_dict('records')
    html = dashboard.build(recs, meta)
    for fname in (f'儀表板_{run_tag}.html', 'dashboard.html'):
        with open(os.path.join(BASE, 'output', fname), 'w', encoding='utf-8') as f:
            f.write(html)

    print(f'\n輸出：{out_path}（{len(out)} 個 SKU）')
    print(f'儀表板：{os.path.join(BASE, "output", "dashboard.html")}（直接雙擊開啟）')
    print('\n方向分布：')
    print(out['方向'].value_counts().to_string())


if __name__ == '__main__':
    main()
